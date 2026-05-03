"""
Script 3: Analysis
================================
Computes hallucination rates and statistics from the experimental data.

USAGE:
    Run after script_1 and script_2 are complete.
    python3 script_3_analysis.py

WHAT IT DOES:
    - Computes overall hallucination rate by demographic group
    - Breaks down by hallucination type (factual/reasoning/evidence)
    - Computes intersectional patterns (Race × Gender)
    - Compares neutral vs sensitive question sets
    - Identifies "worst examples" for case study slide
    - Saves all results to a new Analysis sheet in the Excel file

ESTIMATED TIME: < 1 minute (no API calls)
"""

import sys
from pathlib import Path
from collections import defaultdict

def install_dependencies():
    try:
        import openpyxl
        import pandas
    except ImportError:
        print("Installing required packages...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install",
                               "--quiet", "openpyxl", "pandas"])

install_dependencies()

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_FILE = "medical_llm_disparity_experiments.xlsx"

if not Path(EXCEL_FILE).exists():
    print(f"ERROR: {EXCEL_FILE} not found.")
    sys.exit(1)

# ============================================================
# Load data
# ============================================================
print("="*60)
print("  Analysis Script")
print("="*60 + "\n")

print("📂 Loading data...")
df = pd.read_excel(EXCEL_FILE, sheet_name="Experiments")

# Filter out rows that haven't been judged
judged = df[df["judge_overall"].notna()].copy()
print(f"   Loaded {len(df)} rows total, {len(judged)} judged.\n")

if len(judged) == 0:
    print("⚠️  No judged rows found. Run script_2_judge.py first.")
    sys.exit(0)

# Make sure types are right
for col in ["judge_factual", "judge_reasoning", "judge_evidence", "judge_overall"]:
    judged[col] = pd.to_numeric(judged[col], errors="coerce").fillna(0).astype(int)

# Create combined demographic label
judged["demographic"] = judged["race"] + " " + judged["gender"]

# ============================================================
# Analysis 1: Overall hallucination rate by demographic
# ============================================================
print("📊 ANALYSIS 1: Hallucination Rate by Demographic")
print("-" * 60)

overall = judged.groupby("demographic").agg(
    total=("judge_overall", "count"),
    hallucinations=("judge_overall", "sum"),
    factual=("judge_factual", "sum"),
    reasoning=("judge_reasoning", "sum"),
    evidence=("judge_evidence", "sum"),
).reset_index()

overall["overall_rate"] = (overall["hallucinations"] / overall["total"] * 100).round(1)
overall["factual_rate"] = (overall["factual"] / overall["total"] * 100).round(1)
overall["reasoning_rate"] = (overall["reasoning"] / overall["total"] * 100).round(1)
overall["evidence_rate"] = (overall["evidence"] / overall["total"] * 100).round(1)

print(overall.to_string(index=False))
print()

# ============================================================
# Analysis 2: Neutral vs Sensitive comparison
# ============================================================
print("\n📊 ANALYSIS 2: Neutral vs Sensitive Question Sets")
print("-" * 60)

by_set = judged.groupby(["set_type", "demographic"]).agg(
    total=("judge_overall", "count"),
    hallucinations=("judge_overall", "sum"),
).reset_index()
by_set["rate"] = (by_set["hallucinations"] / by_set["total"] * 100).round(1)

# Pivot for readability
pivot = by_set.pivot(index="demographic", columns="set_type", values="rate").reset_index()
pivot["disparity_neutral_only"] = pivot.get("neutral", 0)
print(pivot.to_string(index=False))
print()

# ============================================================
# Analysis 3: Hallucination Type Breakdown
# ============================================================
print("\n📊 ANALYSIS 3: Hallucination Type by Demographic (Neutral set only)")
print("-" * 60)
print("(Neutral set is most informative — these are pure bias signals)\n")

neutral_only = judged[judged["set_type"] == "neutral"]
type_breakdown = neutral_only.groupby("demographic").agg(
    n=("judge_overall", "count"),
    factual_count=("judge_factual", "sum"),
    reasoning_count=("judge_reasoning", "sum"),
    evidence_count=("judge_evidence", "sum"),
).reset_index()
type_breakdown["factual_rate"] = (type_breakdown["factual_count"] / type_breakdown["n"] * 100).round(1)
type_breakdown["reasoning_rate"] = (type_breakdown["reasoning_count"] / type_breakdown["n"] * 100).round(1)
type_breakdown["evidence_rate"] = (type_breakdown["evidence_count"] / type_breakdown["n"] * 100).round(1)

print(type_breakdown[["demographic", "n", "factual_rate", "reasoning_rate", "evidence_rate"]].to_string(index=False))
print()

# ============================================================
# Analysis 4: Per-question disparity (find dramatic cases)
# ============================================================
print("\n📊 ANALYSIS 4: Per-Question Disparity (Top 5 most asymmetric)")
print("-" * 60)

per_question = judged.groupby(["question_id", "topic", "demographic"]).agg(
    hallucinated=("judge_overall", "sum")
).reset_index()

# Pivot to see all 4 demographics side-by-side per question
q_pivot = per_question.pivot(index=["question_id", "topic"], columns="demographic", values="hallucinated").reset_index()
q_pivot = q_pivot.fillna(0)

# Compute disparity score (range across demographics)
demo_cols = [c for c in q_pivot.columns if c not in ["question_id", "topic"]]
if len(demo_cols) > 0:
    q_pivot["disparity"] = q_pivot[demo_cols].max(axis=1) - q_pivot[demo_cols].min(axis=1)
    q_pivot_sorted = q_pivot.sort_values("disparity", ascending=False)
    print(q_pivot_sorted.head(5).to_string(index=False))
print()

# ============================================================
# Analysis 5: Find "worst example" for case study slide
# ============================================================
print("\n📊 ANALYSIS 5: Best Case Study Candidates")
print("-" * 60)
print("(Questions where one demographic was hallucinated but another was clean)\n")

# Look for questions where Black Female has hallucination but White Male doesn't
case_candidates = []
for qid in judged["question_id"].unique():
    q_subset = judged[judged["question_id"] == qid]
    
    bf = q_subset[(q_subset["race"] == "Black") & (q_subset["gender"] == "woman")]
    wm = q_subset[(q_subset["race"] == "White") & (q_subset["gender"] == "man")]
    
    if len(bf) > 0 and len(wm) > 0:
        bf_hall = bf["judge_overall"].iloc[0]
        wm_hall = wm["judge_overall"].iloc[0]
        if bf_hall == 1 and wm_hall == 0:
            case_candidates.append({
                "question_id": qid,
                "topic": q_subset["topic"].iloc[0],
                "set_type": q_subset["set_type"].iloc[0],
                "issue": "BF hallucinated, WM didn't",
                "bf_notes": str(bf["judge_notes"].iloc[0])[:150] if pd.notna(bf["judge_notes"].iloc[0]) else ""
            })

if case_candidates:
    print(f"Found {len(case_candidates)} dramatic cases:\n")
    for c in case_candidates[:5]:
        print(f"  • {c['question_id']} ({c['topic']}) — {c['issue']}")
        if c['bf_notes']:
            print(f"    Note: {c['bf_notes']}")
        print()
else:
    print("⚠️  No clear-cut Black-Female-vs-White-Male disparity found.")
    print("   Look at Analysis 4 for other interesting patterns.\n")

# ============================================================
# Save analysis to Excel
# ============================================================
print("\n💾 Saving analysis to Excel...")

wb = openpyxl.load_workbook(EXCEL_FILE)

# Remove existing analysis sheet if it exists
if "Analysis" in wb.sheetnames:
    del wb["Analysis"]

ws = wb.create_sheet("Analysis")

header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
header_fill = PatternFill("solid", start_color="2F5597")
section_font = Font(bold=True, name="Arial", size=12, color="2F5597")

current_row = 1

def write_section_header(title):
    global current_row
    cell = ws.cell(current_row, 1, value=title)
    cell.font = section_font
    current_row += 2

def write_dataframe(df_to_write):
    global current_row
    # Headers
    for col_idx, col_name in enumerate(df_to_write.columns, 1):
        cell = ws.cell(current_row, col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    current_row += 1
    # Data
    for _, row in df_to_write.iterrows():
        for col_idx, col_name in enumerate(df_to_write.columns, 1):
            ws.cell(current_row, col_idx, value=row[col_name])
        current_row += 1
    current_row += 2  # spacing

write_section_header("Analysis 1: Hallucination Rate by Demographic")
write_dataframe(overall)

write_section_header("Analysis 2: Neutral vs Sensitive Sets")
write_dataframe(pivot)

write_section_header("Analysis 3: Hallucination Type by Demographic (Neutral)")
write_dataframe(type_breakdown[["demographic", "n", "factual_rate", "reasoning_rate", "evidence_rate"]])

write_section_header("Analysis 4: Per-Question Disparity (sorted)")
if len(demo_cols) > 0:
    write_dataframe(q_pivot_sorted)

write_section_header("Analysis 5: Case Study Candidates")
if case_candidates:
    case_df = pd.DataFrame(case_candidates)
    write_dataframe(case_df)
else:
    ws.cell(current_row, 1, value="No clear BF-vs-WM disparity found in this run.")
    current_row += 2

# Auto-size columns
for col_idx in range(1, 12):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = 18

wb.save(EXCEL_FILE)

print(f"\n✅ Analysis saved to '{EXCEL_FILE}' (new 'Analysis' sheet)")
print("\nNext step: Run script_4_charts.py to generate visualization PNG files")
