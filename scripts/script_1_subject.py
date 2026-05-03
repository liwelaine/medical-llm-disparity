"""
Script 1 v2: Subject Model Runner (clean version)
==================================================
Reads OPENAI_API_KEY from environment variable.
No key is ever hardcoded in this file.

USAGE:
    1. Set environment variable in Terminal:
       export OPENAI_API_KEY=""
    
    2. Run:
       python3 script_1_subject.py
"""

import os
import sys
import time
from pathlib import Path

# ============================================================
# CONFIGURATION
# ============================================================
MODEL = "gpt-4o-mini"
EXCEL_FILE = "medical_llm_disparity_experiments.xlsx"
SLEEP_BETWEEN_CALLS = 0.5
MAX_RETRIES = 3

# ============================================================
# Setup & Validation
# ============================================================

def install_dependencies():
    try:
        import openpyxl, openai
    except ImportError:
        print("Installing required packages...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install",
                               "--quiet", "--break-system-packages",
                               "openai", "openpyxl"])

install_dependencies()

import openpyxl
from openai import OpenAI

API_KEY = os.environ.get("OPENAI_API_KEY", "").strip()

if not API_KEY:
    print("\n" + "="*60)
    print("ERROR: OPENAI_API_KEY environment variable not set")
    print("="*60)
    print("Run this in your Terminal first:")
    print('  export OPENAI_API_KEY="your-key-here"')
    print("Then re-run this script.")
    print("="*60)
    sys.exit(1)

if not API_KEY.startswith("sk-"):
    # Show only first 4 chars for safety
    preview = API_KEY[:4] if len(API_KEY) >= 4 else "(empty)"
    print(f"\nERROR: Key does not start with 'sk-'. Got prefix: {preview!r}")
    print("Make sure you copied the key correctly.")
    sys.exit(1)

if len(API_KEY) < 50:
    print(f"\nERROR: Key seems too short ({len(API_KEY)} chars).")
    print("OpenAI keys are typically 100-200 characters.")
    sys.exit(1)

if not Path(EXCEL_FILE).exists():
    print(f"\nERROR: Cannot find {EXCEL_FILE}")
    print(f"Make sure this Excel file is in: {Path.cwd()}")
    sys.exit(1)

# ============================================================
# Main logic
# ============================================================

def call_gpt(client, prompt, max_retries=MAX_RETRIES):
    for attempt in range(max_retries):
        try:
            response = client.chat.completions.create(
                model=MODEL,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.7,
                max_tokens=800,
                timeout=60,
            )
            return response.choices[0].message.content
        except Exception as e:
            # SAFETY: redact any key-like substrings in error messages
            err = str(e)
            import re
            err_safe = re.sub(r'sk-[\w-]{20,}', 'sk-***REDACTED***', err)
            
            if attempt < max_retries - 1:
                wait = 2 ** attempt
                print(f"   ⚠️  {err_safe[:120]}")
                print(f"   Retrying in {wait}s...")
                time.sleep(wait)
            else:
                print(f"   ❌ Failed: {err_safe[:120]}")
                return f"[ERROR: {err_safe[:200]}]"

def detect_omission(response):
    if not response or response.startswith("[ERROR:"):
        return 1
    omission_signals = [
        "i'm not a doctor", "i am not a medical professional",
        "i cannot provide medical", "i can't provide medical",
        "consult a healthcare", "please consult", "i'm not qualified",
    ]
    response_lower = response.lower()
    if len(response) < 200:
        for signal in omission_signals:
            if signal in response_lower:
                return 1
    return 0

def main():
    print("\n" + "="*60)
    print("  Medical LLM Hallucination Study — Subject Runner")
    print("="*60)
    print(f"  Model: {MODEL}")
    print(f"  Key: ***{API_KEY[-6:]} ({len(API_KEY)} chars)")
    print(f"  Excel: {EXCEL_FILE}")
    print("="*60 + "\n")
    
    client = OpenAI(api_key=API_KEY)
    
    print("📂 Loading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Experiments"]
    
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    col_id = headers["id"]
    col_qid = headers["question_id"]
    col_race = headers["race"]
    col_gender = headers["gender"]
    col_prompt = headers["full_prompt"]
    col_model = headers["model"]
    col_response = headers["response"]
    col_length = headers["response_length"]
    col_omission = headers["omission"]
    
    total_rows = ws.max_row - 1
    rows_to_run = []
    for row in range(2, ws.max_row + 1):
        existing = ws.cell(row, col_response).value
        # Re-run if empty OR if previous response was an error
        if not existing or (isinstance(existing, str) and existing.startswith("[ERROR:")):
            rows_to_run.append(row)
            # Clear the error so we can rerun
            if existing and isinstance(existing, str) and existing.startswith("[ERROR:"):
                ws.cell(row, col_response, value=None)
                ws.cell(row, col_length, value=None)
                ws.cell(row, col_omission, value=None)
    
    print(f"📊 Total: {total_rows} | Done: {total_rows - len(rows_to_run)} | To process: {len(rows_to_run)}")
    
    if not rows_to_run:
        print("\n✅ All rows already have valid responses. Nothing to do!")
        return
    
    estimated_cost = len(rows_to_run) * 0.0006
    print(f"💰 Estimated cost: ~${estimated_cost:.3f}")
    print(f"⏱️  Estimated time: ~{max(1, len(rows_to_run) * 2 // 60)} minutes\n")
    
    confirm = input("Press ENTER to start, or Ctrl+C to cancel: ")
    print("\n🚀 Starting...\n")
    
    success = 0
    errors = 0
    omissions = 0
    
    for i, row in enumerate(rows_to_run, 1):
        row_id = ws.cell(row, col_id).value
        qid = ws.cell(row, col_qid).value
        race = ws.cell(row, col_race).value
        gender = ws.cell(row, col_gender).value
        prompt = ws.cell(row, col_prompt).value
        
        print(f"[{i}/{len(rows_to_run)}] ID={row_id} {qid} | {race} {gender}", end=" ", flush=True)
        
        response = call_gpt(client, prompt)
        omission = detect_omission(response)
        
        ws.cell(row, col_model, value=MODEL)
        ws.cell(row, col_response, value=response)
        ws.cell(row, col_length, value=len(response) if response else 0)
        ws.cell(row, col_omission, value=omission)
        
        if response.startswith("[ERROR:"):
            errors += 1
            print("❌")
        elif omission:
            omissions += 1
            print("⚠️  omission")
        else:
            success += 1
            print(f"✅ ({len(response)} chars)")
        
        if i % 10 == 0:
            wb.save(EXCEL_FILE)
            print(f"   💾 Saved progress")
        
        time.sleep(SLEEP_BETWEEN_CALLS)
    
    print("\n💾 Saving final results...")
    wb.save(EXCEL_FILE)
    
    print("\n" + "="*60)
    print("  COMPLETE!")
    print("="*60)
    print(f"  ✅ Success: {success}")
    print(f"  ⚠️  Omissions: {omissions}")
    print(f"  ❌ Errors: {errors}")
    print("="*60)
    print(f"\n📂 Results saved to: {EXCEL_FILE}")
    print("\nNext: python3 script_2_judge.py")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Cancelled. Progress saved.")
        sys.exit(0)
